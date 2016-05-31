import json
import logging
import os
import pprint
import sys
import uuid
import loader_util
import ckanapi
import inflection
from openpyxl import load_workbook



def create_id(name):
    return inflection.parameterize(name)


def label_to_value(field, label):
    if not label == None:
        sanitised_label = label.replace(u'\xa0', u' ').strip()

        for choice in field['choices']:
            if choice['label'] == sanitised_label:
                return choice['value']

        return ""

        # raise Exception("Label not found in field", sanitised_label)


# Load schema from somewhere?
def load_schema():
    with open(os.path.join(os.path.dirname(os.path.realpath(__file__)), '..',
                           'ckan/plugins/ckanext-datacatalogue_theme/ckanext/datacatalogue_theme',
                           'schema.json')) as data_file:
        data = json.load(data_file)
        return {field['field_name']: field for field in data['dataset_fields']}


if (len(sys.argv)) != 4:
    print """USAGE:
    python loader.py [serviceLocation] [apikey] [inputXls]"""
    sys.exit()

service_location = sys.argv[1]
api_key = sys.argv[2]
input_xls = sys.argv[3]

service = ckanapi.RemoteCKAN(service_location,
                             user_agent='ckanapiexample/1.0 (+http://example.com/my/website)',
                             apikey=api_key
                             )

print 'Using schema'
schema = load_schema()

print 'Loading Datasets'

workbook = load_workbook(input_xls)
sheet = workbook.active


for row in range(2, sheet.max_row+1):

    organisation = "Unknown"
    if sheet.cell(row=row, column=4).value:
        organisationList = sheet.cell(row=row, column=4).value
    #can be a list of organisations
    organisations = organisationList.split(',');
    
    for org in organisations:
        pprint.pprint(org)

        print 'Creating Organisation'
        try:
            service.action.organization_create(
                name=org.strip().lower().replace(' ', '-'),
                title=org
            )
        except:
            pass

    title = "????"

    if sheet.cell(row=row, column=1).value:
        id = create_id(sheet.cell(row=row, column=1).value)[:100]
        title = sheet.cell(row=row, column=1).value
    else:
        continue
        
    try:
        pprint.pprint( " -- " + organisations[0] + " -- " + organisations[0].lower().replace(' ', '-'))
        pprint.pprint( " -- id -- " + id)
        pprint.pprint( " -- title -- " + title)

        

        if sheet.cell(row=row, column=2).value:
            pprint.pprint( " -- summary -- " + sheet.cell(row=row, column=2).value)
        if sheet.cell(row=row, column=3).value:
            pprint.pprint( " -- description -- " + sheet.cell(row=row, column=3).value)
        pprint.pprint( " -- iaoName -- " + sheet.cell(row=row, column=5).value)
        pprint.pprint( " -- iaoEmail -- " + sheet.cell(row=row, column=6).value)
        if sheet.cell(row=row, column=7).value:
            pprint.pprint( " -- personal? -- " +  label_to_value(schema['contains_personal_information'], sheet.cell(row=row, column=7).value)  )
        if sheet.cell(row=row, column=8).value:
            pprint.pprint( " -- policy -- " + sheet.cell(row=row, column=8).value)
        if sheet.cell(row=row, column=9).value:
            pprint.pprint("-- tags --" + sheet.cell(row=row, column=9).value)


        print '  ' + id
        service.action.package_create(
                name=id,
               title=title,
               owner_org=organisations[0].strip().lower().replace(' ', '-'),
               notes=sheet.cell(row=row, column=2).value,
               description=sheet.cell(row=row, column=3).value,
               information_asset_owner_name=sheet.cell(row=row, column=5).value,
               information_asset_owner_email=sheet.cell(row=row, column=6).value,
               contains_personal_information=label_to_value(schema['contains_personal_information'], sheet.cell(row=row, column=7).value),
               data_use_restrictions=sheet.cell(row=row,column=8).value,
               tag_string=sheet.cell(row=row,column=9).value
        )
	
    except Exception, e:
        logging.error(e, exc_info=True)
        print id + ' failed'
    pass

print 'Finished'


